"""Export del listado general de comprobantes a Excel.

Paridad de columnas con el export "Facturas" de Mi Admin XML: mismos 47
encabezados en el mismo orden. Los códigos (uso CFDI, forma de pago, método,
régimen fiscal, tipo) se traducen a 'CÓDIGO - Descripción' con los catálogos
oficiales del SAT. Las columnas de IVA/retenciones se derivan de los
traslados/retenciones guardados en el catálogo.
"""

from __future__ import annotations

import json
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook

from conxml.catalog.db import Catalogo
from conxml.cfdi.catalogos import (
    FORMA_PAGO,
    REGIMEN_FISCAL,
    TIPO_COMPROBANTE,
    USO_CFDI,
    describir,
)
from conxml.export.comun import (
    ESTILO_ENCABEZADO,
    dec_texto as _dec_texto,
    fecha_texto as _fecha,
    fill_estatus as _fill_estatus,
    numero as _numero,
    tasa_es as _tasa_es,
)

ENCABEZADOS = [
    "Verificado o Asoc",
    "Estado SAT",
    "EsCancelable",
    "EstatusCancelacion",
    "CfdiRelacionados",
    "UUID",
    "Serie",
    "Folio",
    "Version",
    "TipoComprobante",
    "FechaTimbradoXML",
    "FechaEmisionXML",
    "LugarDeExpedicion",
    "RFC Emisor",
    "Nombre Emisor",
    "RegimenFiscal",
    "RFC Receptor",
    "Nombre Receptor",
    "UsoCFDI",
    "RegimenFiscalReceptor",
    "DomicilioFiscalReceptor",
    "FormaDePago",
    "Metodo de Pago",
    "Complementos comprobante",
    "Conceptos",
    "Complementos conceptos",
    "SubTotal",
    "Descuento",
    "Total Trasladados",
    "Total Retenidos",
    "Total",
    "Moneda",
    "IVA Exento Base",
    "IVA Cero Base",
    "IVA 8 Importe",
    "IVA 16 Importe",
    "ISR Retenido",
    "IVA Retenido",
    "IEPS Retenido",
    "Ret ISR 1.25 Importe",
    "Ret IVA 10.6667 Importe",
    "Ret IVA 8 Importe",
    "Ret IVA 6 Importe",
    "Ret IVA 16 Importe",
    "No Certificado SAT",
    "No Certificado Emisor",
    "Archivo XML",
]

def _col(nombre: str) -> int:
    return ENCABEZADOS.index(nombre) + 1


COL_FECHA_TIMBRADO = _col("FechaTimbradoXML")
COL_FECHA_EMISION = _col("FechaEmisionXML")
COL_SUBTOTAL = _col("SubTotal")
COL_DESCUENTO = _col("Descuento")
COL_TRASLADADOS = _col("Total Trasladados")
COL_RETENIDOS = _col("Total Retenidos")
COL_TOTAL = _col("Total")
COL_IVA_EXENTO = _col("IVA Exento Base")
COL_IVA_CERO = _col("IVA Cero Base")
COL_IVA_8 = _col("IVA 8 Importe")
COL_IVA_16 = _col("IVA 16 Importe")
COL_ISR = _col("ISR Retenido")
COL_IVA_RET = _col("IVA Retenido")
COL_IEPS = _col("IEPS Retenido")
COL_RET_ISR_125 = _col("Ret ISR 1.25 Importe")
COL_RET_IVA_106667 = _col("Ret IVA 10.6667 Importe")
COL_RET_IVA_8 = _col("Ret IVA 8 Importe")
COL_RET_IVA_6 = _col("Ret IVA 6 Importe")
COL_RET_IVA_16 = _col("Ret IVA 16 Importe")
COL_ESTADO_SAT = _col("Estado SAT")

COMPLEMENTOS_IGNORADOS = {"TimbreFiscalDigital"}

ANCHOS = {
    "A": 10, "B": 13, "C": 22, "D": 20, "E": 38, "F": 38, "G": 8, "H": 10,
    "I": 8, "J": 15, "K": 19, "L": 19, "M": 15, "N": 15, "O": 32, "P": 32,
    "Q": 15, "R": 32, "S": 22, "T": 32, "U": 20, "V": 26, "W": 32, "X": 18,
    "Y": 40, "Z": 18, "AA": 14, "AB": 14, "AC": 14, "AD": 14, "AE": 14,
    "AF": 16, "AG": 16, "AH": 16, "AI": 16, "AJ": 16, "AK": 18, "AL": 18,
    "AM": 18, "AN": 18, "AO": 18, "AP": 18, "AQ": 18, "AR": 18, "AS": 18,
    "AT": 18, "AU": 40,
}


def _fecha(valor: str | None):  # compat: ahora en comun
    from conxml.export.comun import fecha_texto

    return fecha_texto(valor)


def _dec_texto(valor: str | None):  # compat: ahora en comun
    from conxml.export.comun import dec_texto

    return dec_texto(valor)


def _importes(json_texto: str | None) -> list[dict]:
    if not json_texto:
        return []
    try:
        data = json.loads(json_texto)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def _tasa_es(valor: str | None, tasa: str):  # compat: ahora en comun
    from conxml.export.comun import tasa_es

    return tasa_es(valor, tasa)


def _traslado_por_tasa(traslados: list[dict], tasa: str) -> float | None:
    return _dec_texto(
        next(
            (
                t["importe"]
                for t in traslados
                if t.get("impuesto") == "002"
                and t.get("tipo_factor") == "Tasa"
                and _tasa_es(t.get("tasa_o_cuota"), tasa)
                and t.get("importe") is not None
            ),
            None,
        )
    )


def _base_por_factor(traslados: list[dict], factor: str) -> float | None:
    return _dec_texto(
        next(
            (
                t["base"]
                for t in traslados
                if t.get("impuesto") == "002"
                and t.get("tipo_factor") == factor
                and t.get("base") is not None
            ),
            None,
        )
    )


def _retencion_por_impuesto(retenciones: list[dict], impuesto: str) -> float | None:
    return _dec_texto(
        next(
            (
                r["importe"]
                for r in retenciones
                if r.get("impuesto") == impuesto and r.get("importe") is not None
            ),
            None,
        )
    )


def _retencion_por_tasa(retenciones: list[dict], impuesto: str, tasa: str) -> float | None:
    return _dec_texto(
        next(
            (
                r["importe"]
                for r in retenciones
                if r.get("impuesto") == impuesto
                and _tasa_es(r.get("tasa_o_cuota"), tasa)
                and r.get("importe") is not None
            ),
            None,
        )
    )


def _suma_importes(lista: list[dict]) -> float | None:
    totales: list[Decimal] = []
    for t in lista:
        if t.get("importe") is None:
            continue
        try:
            totales.append(Decimal(t["importe"]))
        except (InvalidOperation, ValueError, TypeError):
            continue
    if not totales:
        return None
    return float(sum(totales, Decimal(0)).quantize(Decimal("0.01")))


def exportar_listado(
    catalogo: Catalogo,
    destino: str | Path,
    cliente: str | None = None,
    desde: str | None = None,
    hasta: str | None = None,
) -> Path:
    """Genera un Excel con el listado general (47 columnas, paridad Mi Admin XML)."""
    destino = Path(destino)
    filas = catalogo.consulta(cliente=cliente, desde=desde, hasta=hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Listado"
    ws.append(ENCABEZADOS)
    for celda in ws[1]:
        celda.font = ESTILO_ENCABEZADO

    for fila in filas:
        traslados = _importes(fila["traslados_json"])
        retenciones = _importes(fila["retenciones_json"])
        complementos = [
            c.strip()
            for c in (fila["complementos"] or "").replace(";", ",").split(",")
            if c.strip() and c.strip() not in COMPLEMENTOS_IGNORADOS
        ]

        ws.append(
            [
                None,  # Verificado o Asoc
                fila["estatus"],
                fila["es_cancelable"],
                fila["estatus_cancelacion"],
                fila["relaciones"],
                fila["uuid"],
                fila["serie"],
                fila["folio"],
                fila["version"],
                describir(TIPO_COMPROBANTE, fila["tipo_comprobante"]),
                _fecha(fila["fecha_timbrado"]),
                _fecha(fila["fecha"]),
                fila["lugar_expedicion"],
                fila["emisor_rfc"],
                fila["emisor_nombre"],
                describir(REGIMEN_FISCAL, fila["emisor_regimen_fiscal"]),
                fila["receptor_rfc"],
                fila["receptor_nombre"],
                describir(USO_CFDI, fila["uso_cfdi"]),
                describir(REGIMEN_FISCAL, fila["regimen_fiscal_receptor"]),
                fila["domicilio_fiscal_receptor"],
                describir(FORMA_PAGO, fila["forma_pago"]),
                fila["metodo_pago"],  # Mi Admin XML deja el código crudo
                ", ".join(complementos) or None,
                fila["conceptos"],
                None,  # Complementos conceptos
                _numero(fila["subtotal"]),
                _numero(fila["descuento"]),
                _suma_importes(traslados),
                _suma_importes(retenciones),
                _numero(fila["total"]),
                fila["moneda"],
                _base_por_factor(traslados, "Exento"),
                _base_por_factor(traslados, "Tasa")
                if any(
                    _tasa_es(t.get("tasa_o_cuota"), "0")
                    for t in traslados
                    if t.get("impuesto") == "002" and t.get("tipo_factor") == "Tasa"
                )
                else None,
                _traslado_por_tasa(traslados, "0.08"),
                _traslado_por_tasa(traslados, "0.16"),
                _retencion_por_impuesto(retenciones, "001"),
                _retencion_por_impuesto(retenciones, "002"),
                _retencion_por_impuesto(retenciones, "003"),
                _retencion_por_tasa(retenciones, "001", "0.0125"),
                _retencion_por_tasa(retenciones, "002", "0.106667"),
                _retencion_por_tasa(retenciones, "002", "0.08"),
                _retencion_por_tasa(retenciones, "002", "0.06"),
                _retencion_por_tasa(retenciones, "002", "0.16"),
                fila["no_certificado_sat"],
                fila["no_certificado_emisor"],
                Path(fila["ruta"]).name if fila["ruta"] else None,
            ]
        )
        n = ws.max_row
        for col in (
            COL_SUBTOTAL,
            COL_DESCUENTO,
            COL_TRASLADADOS,
            COL_RETENIDOS,
            COL_TOTAL,
            COL_IVA_EXENTO,
            COL_IVA_CERO,
            COL_IVA_8,
            COL_IVA_16,
            COL_ISR,
            COL_IVA_RET,
            COL_IEPS,
            COL_RET_ISR_125,
            COL_RET_IVA_106667,
            COL_RET_IVA_8,
            COL_RET_IVA_6,
            COL_RET_IVA_16,
        ):
            ws.cell(row=n, column=col).number_format = "#,##0.00"
        fill = _fill_estatus(fila["estatus"])
        if fill is not None:
            ws.cell(row=n, column=COL_ESTADO_SAT).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino