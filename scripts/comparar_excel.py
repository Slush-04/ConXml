"""Comparación columna a columna: Mi Admin XML vs ConXml (47 columnas).

Cruza por UUID y compara los valores de cada columna en las dos hojas.
Ignora diferencias puramente de formato (None == "" ; ".." == "").
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

A = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/comparacion/miadm.xlsx")
B = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/comparacion/conxml.xlsx")

COMPARABLES = {
    "Estado SAT": "Estado SAT",
    "EsCancelable": "EsCancelable",
    "EstatusCancelacion": "EstatusCancelacion",
    "UUID": "UUID",
    "Serie": "Serie",
    "Folio": "Folio",
    "Version": "Version",
    "TipoComprobante": "TipoComprobante",
    "FechaTimbradoXML": "FechaTimbradoXML",
    "FechaEmisionXML": "FechaEmisionXML",
    "LugarDeExpedicion": "LugarDeExpedicion",
    "RFC Emisor": "RFC Emisor",
    "Nombre Emisor": "Nombre Emisor",
    "RegimenFiscal": "RegimenFiscal",
    "RFC Receptor": "RFC Receptor",
    "Nombre Receptor": "Nombre Receptor",
    "UsoCFDI": "UsoCFDI",
    "FormaDePago": "FormaDePago",
    "Metodo de Pago": "Metodo de Pago",
    "SubTotal": "SubTotal",
    "Descuento": "Descuento",
    "Total Trasladados": "Total Trasladados",
    "Total Retenidos": "Total Retenidos",
    "Total": "Total",
    "Moneda": "Moneda",
    "IVA Exento Base": "IVA Exento Base",
    "IVA Cero Base": "IVA Cero Base",
    "IVA 8 Importe": "IVA 8 Importe",
    "IVA 16 Importe": "IVA 16 Importe",
    "ISR Retenido": "ISR Retenido",
    "IVA Retenido": "IVA Retenido",
    "IEPS Retenido": "IEPS Retenido",
    "No Certificado SAT": "No Certificado SAT",
    "No Certificado Emisor": "No Certificado Emisor",
}


def leer(path: Path) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    enc = [str(c).strip() if c is not None else "" for c in filas[0]]
    return enc, [dict(zip(enc, fila)) for fila in filas[1:] if any(fila)]


def norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main() -> int:
    enc_a, filas_a = leer(A)
    enc_b, filas_b = leer(B)

    print(f"Mi Admin XML: {len(filas_a)} filas x {len(enc_a)} cols | ConXml: {len(filas_b)} filas x {len(enc_b)} cols")
    print(f"Encabezados MiAdmin: {len(enc_a)} | ConXml: {len(enc_b)}")

    faltantes_b = [c for c in enc_a if c not in enc_b]
    extra_b = [c for c in enc_b if c not in enc_a]
    print(f"\nColumnas MiAdmin que faltan en ConXml: {faltantes_b or 'NINGUNA'}")
    print(f"Columnas extra en ConXml: {extra_b or 'NINGUNA'}")

    ma = {norm(f.get("UUID")): f for f in filas_a}
    cb = {norm(f.get("UUID")): f for f in filas_b}
    comunes = sorted(set(ma) & set(cb))
    print(f"\nUUIDs comunes: {len(comunes)} | solo MiAdmin: {len(set(ma) - set(cb))} | solo ConXml: {len(set(cb) - set(ma))}")

    diffs = 0
    for uuid in comunes:
        f_a, f_b = ma[uuid], cb[uuid]
        campos_diff = []
        for campo_xml, campo_cx in COMPARABLES.items():
            va, vb = norm(f_a.get(campo_xml)), norm(f_b.get(campo_cx))
            if va != vb:
                diffs += 1
                campos_diff.append(f"{campo_xml}: MiAdmin={va!r} vs ConXml={vb!r}")
        if campos_diff:
            print(f"  DIFF {uuid} (folio {f_a.get('Folio')})")
            for c in campos_diff:
                print(f"      {c}")

    total_a = sum(float(f.get("Total") or 0) for f in filas_a)
    total_b = sum(float(f.get("Total") or 0) for f in filas_b)
    print(f"\nTotal: MiAdmin=${total_a:,.2f} | ConXml=${total_b:,.2f} | diff=${total_a - total_b:,.2f}")
    print(f"Diferencias encontradas: {diffs}")
    return 0


if __name__ == "__main__":
    sys.exit(main())