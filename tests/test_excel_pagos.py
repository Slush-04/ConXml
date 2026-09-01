"""Tests del export de conciliación de pagos a Excel."""

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from conxml.catalog.db import Catalogo
from conxml.catalog.importer import importar_carpeta
from conxml.export.pagos import exportar_pagos

FIXTURES = Path(__file__).parent / "fixtures"


@pytest.fixture()
def catalogo(tmp_path):
    lib = Catalogo(tmp_path / "catalogo.db")
    carpeta = tmp_path / "cliente1"
    carpeta.mkdir(parents=True, exist_ok=True)
    for nombre in [
        "ingreso_iva.xml",  # uuid 123E4567... (pagado por pago_rep_basico)
        "ingreso_descuento.xml",  # uuid 5A6B7C8D... (pagado por pago_rep_multiple)
        "pago_rep_basico.xml",
        "pago_rep_multiple.xml",
    ]:
        shutil.copy2(FIXTURES / nombre, carpeta / nombre)
    importar_carpeta(lib, carpeta, "1")
    for fila in lib.consulta():
        lib.asignar_estatus(fila["uuid"], "Vigente")
    yield lib
    lib.close()


def test_conciliacion_formato_plano(catalogo, tmp_path):
    destino = exportar_pagos(catalogo, tmp_path / "salida" / "pagos.xlsx")

    wb = load_workbook(destino)
    ws = wb["Pagos"]
    filas = list(ws.iter_rows(values_only=True))
    # 1 docto (basico) + 4 doctos (multiple) = 5 filas + encabezado
    assert len(filas) == 6
    assert filas[0][0] == "Cliente"
    assert filas[0][14] == "UUID Factura"

    fila1 = filas[1]
    assert fila1[0] == "1"  # cliente
    assert fila1[1] == "A1B2C3D4-E5F6-4A7B-8C9D-0E1F2A3B4C5D"  # UUID REP
    assert fila1[6] == "2024-02-01T10:00:00"  # fecha pago
    assert fila1[9] == 1160.00  # monto
    assert fila1[14] == "123E4567-E89B-12D3-A456-426614174000"  # UUID factura
    assert fila1[22] == "Vigente"  # estatus factura (join al catálogo)
    assert fila1[18] == 1  # parcialidad
    assert fila1[21] == 0.00  # saldo insoluto

    # metadatos de paridad (tratamiento del listado general)
    assert fila1[7] == "03 - Transferencia electrónica de fondos"  # forma pago descrita
    assert fila1[27] == "P - Pago"  # tipo comprobante REP descrito
    assert filas[0][24] == "EsCancelable REP"
    assert filas[0][29] == "EsCancelable"
    assert filas[0][33] == "Total Factura"
    factura_descuento = next(
        f for f in filas[1:] if f[14] == "5A6B7C8D-9E0F-4A1B-8C2D-3E4F5A6B7C8D"
    )
    assert factura_descuento[31] == "2024-02-10T14:35:00"  # fecha timbrado factura
    assert factura_descuento[32] == "I - Ingreso"  # tipo comprobante factura descrito
    assert factura_descuento[33] == 1044.00  # total de la factura pagada

    # parcialidades del REP múltiple: saldos altos -> pendientes
    pendiente = [f for f in filas[1:] if f[18] == 1 and f[21] > 0][0]
    assert pendiente[21] == 300.00
    celda = ws.cell(row=filas.index(pendiente) + 1, column=22)
    assert celda.fill.start_color.rgb == "00FFEB9C"  # amarillo pendiente


def test_resumen_por_cliente_y_moneda(catalogo, tmp_path):
    destino = exportar_pagos(catalogo, tmp_path / "pagos2.xlsx")

    wb = load_workbook(destino)
    ws = wb["Resumen"]
    filas = list(ws.iter_rows(values_only=True))
    assert filas[0][0] == "Cliente"
    por_moneda = {f[1]: f for f in filas[1:]}
    assert por_moneda["MXN"][2] == 2  # pagos MXN: 1160 + 20000
    assert por_moneda["MXN"][3] == 21160.00
    assert por_moneda["MXN"][4] == 1  # 1 factura con saldo insoluto (38000)
    assert por_moneda["USD"][2] == 1
    assert por_moneda["USD"][3] == 300.00


def test_conciliacion_sin_pagos_ok(catalogo, tmp_path):
    destino = exportar_pagos(catalogo, tmp_path / "pagos3.xlsx")
    wb = load_workbook(destino)
    assert list(wb["Pagos"].iter_rows(values_only=True))[0][0] == "Cliente"