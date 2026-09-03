"""Helpers compartidos de exportación Excel (antes triplicados).

Centraliza ``_numero/_fecha/_fill_estatus``, estilos y anchos para
listado/pagos/nómina. Los módulos viejos siguen funcionando porque
re-exportan estos nombres por compatibilidad.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl.styles import Font, PatternFill

ESTILO_ENCABEZADO = Font(bold=True)
FILL_VIGENTE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_CANCELADO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_SIN_DATO = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_PENDIENTE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def fill_estatus(estatus: str | None) -> PatternFill | None:
    match estatus:
        case "Vigente":
            return FILL_VIGENTE
        case "Cancelado":
            return FILL_CANCELADO
        case None | "Sin validar":
            return FILL_SIN_DATO
        case _:
            return None


def numero(valor: str | float | Decimal | None) -> float | None:
    """Tolerante a None/'' (antes float('') lanzaba ValueError)."""
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (ValueError, TypeError):
        return None


def dec_texto(valor: str | None) -> float | None:
    if not valor:
        return None
    try:
        return float(Decimal(valor))
    except (InvalidOperation, ValueError, TypeError):
        return None


def fecha_texto(valor: str | None) -> str | None:
    """Fecha como texto ISO '2026-07-14T09:48:42' (formato Mi Admin XML)."""
    if not valor:
        return None
    try:
        dt = datetime.fromisoformat(valor)
    except ValueError:
        return valor
    return dt.isoformat(timespec="seconds")


def tasa_es(valor: str | None, tasa: str) -> bool:
    if not valor:
        return False
    try:
        return Decimal(valor) == Decimal(tasa)
    except (ValueError, InvalidOperation):
        return False


# Aliases de compatibilidad con los imports privados antiguos.
_fill_estatus = fill_estatus
_numero = numero
_fecha = fecha_texto
_dec_texto = dec_texto
_tasa_es = tasa_es

__all__ = [
    "ESTILO_ENCABEZADO",
    "FILL_VIGENTE",
    "FILL_CANCELADO",
    "FILL_SIN_DATO",
    "FILL_PENDIENTE",
    "fill_estatus",
    "numero",
    "dec_texto",
    "fecha_texto",
    "tasa_es",
]
