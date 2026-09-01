"""Tests del export del listado general a Excel (47 columnas, paridad Mi Admin XML)."""

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from conxml.catalog.db import Catalogo
from conxml.catalog.importer import importar_carpeta
from conxml.export.listado import ENCABEZADOS, exportar_listado

FIXTURES = Path(__file__).parent / "fixtures"

COL_ESTADO_SAT = 2
COL_UUID = 6
COL_TIPO = 10
COL_SUBTOTAL = 27
COL_IVA = 30  # placeholder, se ajusta debajo
COL_TOTAL = 31
COL_FECHA_EMISION = 12


@pytest.fixture()
def catalogo(tmp_path):
    lib = Catalogo(tmp_path / "catalogo.db")
    carpeta = tmp_path / "cliente1"
    carpeta.mkdir(parents=True, exist_ok=True)
    for nombre in [
        "ingreso_iva.xml",
        "ingreso_descuento.xml",
        "egreso_nota_credito.xml",
    ]:
        shutil.copy2(FIXTURES / nombre, carpeta / nombre)
    importar_carpeta(lib, carpeta, "1")
    for fila in lib.consulta():
        estado = "Cancelado" if fila["tipo_comprobante"] == "E" else "Vigente"
        lib.asignar_estatus(fila["uuid"], estado)
    yield lib
    lib.close()


def test_listado_47_columnas_paridad_miadmin(catalogo, tmp_path):
    destino = exportar_listado(catalogo, tmp_path / "salida" / "listado.xlsx")

    wb = load_workbook(destino)
    ws = wb["Listado"]

    filas = list(ws.iter_rows(values_only=True))
    assert len(filas) == 4  # encabezado + 3 comprobantes
    assert len(ENCABEZADOS) == 47

    enc = list(filas[0])
    assert enc[0] == "Verificado o Asoc"
    assert enc[1] == "Estado SAT"
    assert enc[30] == "Total"
    assert enc[46] == "Archivo XML"
    assert len(enc) == 47

    ingreso = [f for f in filas[1:] if f[9] == "I - Ingreso"][0]
    assert ingreso[0] is None  # Verificado o Asoc vacío
    assert ingreso[1] == "Vigente"

    con_iva = [f for f in filas[1:] if f[9] == "I - Ingreso" and f[5] == "123E4567-E89B-12D3-A456-426614174000"][0]
    assert con_iva[26] == 1000.00  # subtotal numérico
    assert con_iva[35] == 160.00  # IVA 16 Importe
    assert con_iva[30] == 1160.00  # total

    con_descuento = [f for f in filas[1:] if f[9] == "I - Ingreso" and f[5] != "123E4567-E89B-12D3-A456-426614174000"][0]
    assert con_descuento[27] == 100.00  # descuento
    assert con_descuento[30] == 1044.00  # total
    assert con_descuento[35] == 144.00  # IVA 16 sobre base 900

    egreso = [f for f in filas[1:] if f[9] == "E - Egreso"][0]
    assert egreso[1] == "Cancelado"

    # la fila de egreso relacionada trae los UUIDs de las facturas en CfdiRelacionados
    assert egreso[4].startswith("123E4567-E89B-12D3-A456-426614174000, 5A6B7C8D")

    # colores de la celda de estatus
    for row in range(2, ws.max_row + 1):
        celda = ws.cell(row=row, column=2)
        if celda.value == "Vigente":
            assert celda.fill.start_color.rgb == "00C6EFCE"
        elif celda.value == "Cancelado":
            assert celda.fill.start_color.rgb == "00FFC7CE"

    # fechas en texto ISO con T (formato de Mi Admin XML)
    assert con_iva[11] == "2024-02-01T09:15:00"


def test_listado_filtro_por_cliente(tmp_path):
    catalogo = Catalogo(tmp_path / "catalogo.db")
    c1 = tmp_path / "c1"
    c1.mkdir(parents=True, exist_ok=True)
    shutil.copy2(FIXTURES / "ingreso_iva.xml", c1 / "ingreso_iva.xml")
    importar_carpeta(catalogo, c1, "1")
    c2 = tmp_path / "c2"
    c2.mkdir(parents=True, exist_ok=True)
    shutil.copy2(FIXTURES / "ingreso_descuento.xml", c2 / "ingreso_descuento.xml")
    importar_carpeta(catalogo, c2, "2")

    destino = exportar_listado(catalogo, tmp_path / "listado2.xlsx", cliente="2")
    wb = load_workbook(destino)
    ws = wb["Listado"]
    filas = list(ws.iter_rows(values_only=True))
    assert len(filas) == 2  # encabezado + 1
    assert all(f[9] == "I - Ingreso" for f in filas[1:])  # solo el cliente 2

    # sin estatus consultado: celda gris
    assert ws.cell(row=2, column=COL_ESTADO_SAT).fill.start_color.rgb == "00F2F2F2"