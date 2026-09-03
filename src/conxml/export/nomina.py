"""Export de recibos de nómina a Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from conxml.catalog.db import Catalogo
from conxml.cfdi import parse_comprobante, parse_nomina
from conxml.export.comun import (
    ESTILO_ENCABEZADO,
    fecha_texto as _fecha,
    fill_estatus as _fill_estatus,
    numero as _numero,
)

ENCABEZADOS_NOMINA = [
    "Cliente",
    "Estado SAT",
    "EsCancelable",
    "EstatusCancelacion",
    "UUID",
    "Serie",
    "Folio",
    "FechaTimbradoXML",
    "FechaEmisionXML",
    "RFC Emisor",
    "Nombre Emisor",
    "Registro Patronal",
    "RFC Receptor",
    "Nombre Receptor",
    "CURP",
    "Num Empleado",
    "Puesto",
    "Tipo Nomina",
    "Fecha Pago",
    "Fecha Inicial Pago",
    "Fecha Final Pago",
    "Num Dias Pagados",
    "Periodicidad Pago",
    "Total Percepciones",
    "Total Deducciones",
    "Total Otros Pagos",
    "SubTotal",
    "Descuento",
    "Total",
    "Moneda",
    "Archivo XML",
]

COL_ESTADO_SAT = 2
COL_SUBTOTAL = 27
COL_DESCUENTO = 28
COL_TOTAL = 29


def _a_decimal(valor) -> float | None:
    from decimal import Decimal

    if valor is None or valor == "":
        return None
    try:
        return float(Decimal(valor))
    except Exception:
        return None


def exportar_nomina(
    catalogo: Catalogo,
    destino: str | Path,
    cliente: str | None = None,
) -> Path:
    """Genera un Excel con los recibos de Nómina del catálogo."""
    import logging

    log = logging.getLogger("conxml.export.nomina")
    destino = Path(destino)
    filas = list(catalogo.consulta(cliente=cliente, tipo="N"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"
    ws.append(ENCABEZADOS_NOMINA)
    for celda in ws[1]:
        celda.font = ESTILO_ENCABEZADO

    for fila in filas:
        path_xml = Path(fila["ruta"])
        nom_data = None
        if path_xml.is_file():
            try:
                comp = parse_comprobante(path_xml)
                nom_data = parse_nomina(comp)
            except Exception as exc:  # noqa: BLE001 — no tumbar export por 1 XML
                log.warning("nómina %s: %s", fila["ruta"], exc)

        reg_patronal = nom_data.registro_patronal if nom_data else None
        curp = nom_data.curp if nom_data else None
        num_emp = nom_data.num_empleado if nom_data else None
        puesto = nom_data.puesto if nom_data else None
        tipo_nom = nom_data.tipo_nomina if nom_data else None
        f_pago = nom_data.fecha_pago.strftime("%Y-%m-%d") if (nom_data and nom_data.fecha_pago) else None
        f_ini = nom_data.fecha_inicial_pago.strftime("%Y-%m-%d") if (nom_data and nom_data.fecha_inicial_pago) else None
        f_fin = nom_data.fecha_final_pago.strftime("%Y-%m-%d") if (nom_data and nom_data.fecha_final_pago) else None
        dias_pag = _a_decimal(nom_data.num_dias_pagados) if nom_data else None
        period = nom_data.periodicidad_pago if nom_data else None
        t_per = _a_decimal(nom_data.total_percepciones) if nom_data else None
        t_ded = _a_decimal(nom_data.total_deducciones) if nom_data else None
        t_otr = _a_decimal(nom_data.total_otros_pagos) if nom_data else None

        ws.append([
            fila["cliente"],
            fila["estatus"],
            fila["es_cancelable"],
            fila["estatus_cancelacion"],
            fila["uuid"],
            fila["serie"],
            fila["folio"],
            _fecha(fila["fecha_timbrado"]),
            _fecha(fila["fecha"]),
            fila["emisor_rfc"],
            fila["emisor_nombre"],
            reg_patronal,
            fila["receptor_rfc"],
            fila["receptor_nombre"],
            curp,
            num_emp,
            puesto,
            tipo_nom,
            f_pago,
            f_ini,
            f_fin,
            dias_pag,
            period,
            t_per,
            t_ded,
            t_otr,
            _numero(fila["subtotal"]),
            _numero(fila["descuento"]),
            _numero(fila["total"]),
            fila["moneda"],
            Path(fila["ruta"]).name if fila["ruta"] else None,
        ])
        n = ws.max_row
        for col in (COL_SUBTOTAL, COL_DESCUENTO, COL_TOTAL):
            ws.cell(row=n, column=col).number_format = "#,##0.00"
        fill = _fill_estatus(fila["estatus"])
        if fill is not None:
            ws.cell(row=n, column=COL_ESTADO_SAT).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino
