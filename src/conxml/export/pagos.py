"""Export de conciliación de pagos (REP) a Excel, formato plano.

Una fila por (pago x documento relacionado). Incluye, además de los datos
del REP, la conciliación: parcialidad, saldos y el estatus SAT de cada
factura pagada (vía join al catálogo), con el mismo tratamiento de
metadatos que el listado general (descripciones SAT, EsCancelable,
EstatusCancelacion, FechaTimbradoXML, Archivo XML).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from conxml.catalog.db import Catalogo
from conxml.cfdi.catalogos import FORMA_PAGO, TIPO_COMPROBANTE, describir
from conxml.export.comun import (
    ESTILO_ENCABEZADO,
    FILL_PENDIENTE,
    fecha_texto as _fecha,
    fill_estatus as _fill_estatus,
    numero as _numero,
)

ENCABEZADOS = [
    "Cliente",
    "UUID REP",
    "Serie REP",
    "Folio REP",
    "Fecha REP",
    "Estatus REP",
    "Fecha Pago",
    "Forma Pago",
    "Moneda Pago",
    "Monto",
    "Núm. Operación",
    "Cuenta Origen",
    "Cuenta Destino",
    "RFC Cta Origen",
    "UUID Factura",
    "Serie",
    "Folio",
    "Moneda DR",
    "Parcialidad",
    "Saldo Anterior",
    "Pagado",
    "Saldo Insoluto",
    "Estatus Factura",
    "Estatus consultado",
    "EsCancelable REP",
    "EstatusCancelacion REP",
    "FechaTimbrado REP",
    "TipoComprobante REP",
    "Archivo XML REP",
    "EsCancelable",
    "EstatusCancelacion",
    "FechaTimbradoXML",
    "TipoComprobante",
    "Total Factura",
    "Archivo XML",
]

COL_SALDO_INSOLUTO = 22
COL_ESTATUS_FACTURA = 23
COL_ESTATUS_REP = 6
COL_FECHA_REP = 5
COL_FECHA_PAGO = 7
COL_MONTO = 10
COL_PAGADO = 21
COL_INSOLUTO = COL_SALDO_INSOLUTO
COL_ES_CANCELABLE_REP = 25
COL_ESTATUS_CANCELACION_REP = 26
COL_FECHA_TIMBRADO_REP = 27
COL_FECHA_TIMBRADO_FACTURA = 32
COL_TOTAL_FACTURA = 34

ANCHOS = {
    "A": 8, "B": 38, "C": 8, "D": 8, "E": 17, "F": 12, "G": 17,
    "H": 34, "I": 12, "J": 12, "K": 13, "L": 13, "M": 13, "N": 13, "O": 38,
    "P": 8, "Q": 8, "R": 10, "S": 12, "T": 13, "U": 12, "V": 13, "W": 15,
    "X": 20, "Y": 13, "Z": 15, "AA": 19, "AB": 30, "AC": 33, "AD": 13,
    "AE": 15, "AF": 19, "AG": 30, "AH": 13, "AI": 33,
}


def _numero(valor):  # compat: ahora en comun
    from conxml.export.comun import numero

    return numero(valor)


def _fill_estatus(estatus):  # compat: ahora en comun
    from conxml.export.comun import fill_estatus

    return fill_estatus(estatus)


def exportar_pagos(
    catalogo: Catalogo,
    destino: str | Path,
    cliente: str | None = None,
) -> Path:
    """Genera un Excel de conciliación de pagos y devuelve la ruta."""
    destino = Path(destino)
    pagos = list(catalogo.consultar_pagos(cliente=cliente))
    comprobantes = {f["uuid"]: f for f in catalogo.consulta(cliente=cliente)}
    todo = {f["uuid"]: f for f in catalogo.consulta()}
    doctos_por_pago = catalogo.consultar_doctos_lote([p["id"] for p in pagos])

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.append(ENCABEZADOS)
    for celda in ws[1]:
        celda.font = ESTILO_ENCABEZADO

    resumen: dict[tuple[str, ...], dict[str, object]] = {}

    for pago in pagos:
        rep = comprobantes.get(pago["comprobante_uuid"])
        clave = (pago["cliente"], pago["moneda"] or "")
        datos = resumen.setdefault(clave, {"pagos": 0, "monto": 0.0, "pendientes": 0, "insoluto": 0.0})
        datos["pagos"] += 1
        datos["monto"] += _numero(pago["monto"]) or 0

        for doc in doctos_por_pago.get(pago["id"], []):
            factura = todo.get(doc["uuid_doc"])
            insoluble = _numero(doc["imp_saldo_insoluto"]) or 0
            if insoluble > 0.01:
                datos["pendientes"] += 1
                datos["insoluto"] += insoluble

            ws.append(
                [
                    pago["cliente"],
                    rep["uuid"] if rep else pago["comprobante_uuid"],
                    rep["serie"] if rep else None,
                    rep["folio"] if rep else None,
                    _fecha(rep["fecha"]) if rep else None,
                    rep["estatus"] if rep else None,
                    pago["fecha_pago"],
                    describir(FORMA_PAGO, pago["forma_pago"]),
                    pago["moneda"],
                    _numero(pago["monto"]),
                    pago["num_operacion"],
                    pago["cta_ordenante"],
                    pago["cta_beneficiario"],
                    pago["rfc_emisor_cta_ord"],
                    doc["uuid_doc"],
                    doc["serie"],
                    doc["folio"],
                    doc["moneda"],
                    doc["num_parcialidad"],
                    _numero(doc["imp_saldo_ant"]),
                    _numero(doc["imp_pagado"]),
                    _numero(doc["imp_saldo_insoluto"]),
                    factura["estatus"] if factura else None,
                    factura["estatus_fecha"] if factura else None,
                    rep["es_cancelable"] if rep else None,
                    rep["estatus_cancelacion"] if rep else None,
                    _fecha(rep["fecha_timbrado"]) if rep else None,
                    describir(TIPO_COMPROBANTE, rep["tipo_comprobante"]) if rep else None,
                    Path(rep["ruta"]).name if rep and rep["ruta"] else None,
                    factura["es_cancelable"] if factura else None,
                    factura["estatus_cancelacion"] if factura else None,
                    _fecha(factura["fecha_timbrado"]) if factura else None,
                    describir(TIPO_COMPROBANTE, factura["tipo_comprobante"]) if factura else None,
                    _numero(factura["total"]) if factura else None,
                    Path(factura["ruta"]).name if factura and factura["ruta"] else None,
                ]
            )
            n = ws.max_row
            ws.cell(row=n, column=COL_FECHA_REP).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row=n, column=COL_FECHA_PAGO).number_format = "yyyy-mm-dd hh:mm"
            for col in (COL_MONTO, COL_PAGADO, COL_INSOLUTO, COL_TOTAL_FACTURA):
                ws.cell(row=n, column=col).number_format = "#,##0.00"
            fill_rep = _fill_estatus(rep["estatus"] if rep else None)
            if fill_rep is not None:
                ws.cell(row=n, column=COL_ESTATUS_REP).fill = fill_rep
            if insoluble > 0.01:
                ws.cell(row=n, column=COL_SALDO_INSOLUTO).fill = FILL_PENDIENTE
            fill = _fill_estatus(factura["estatus"] if factura else None)
            if fill is not None:
                ws.cell(row=n, column=COL_ESTATUS_FACTURA).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for letra, ancho in ANCHOS.items():
        ws.column_dimensions[letra].width = ancho

    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.append(["Cliente", "Moneda", "Pagos", "Monto Total", "Facturas Pendientes", "Saldo Insoluto"])
    for celda in ws_resumen[1]:
        celda.font = ESTILO_ENCABEZADO
    for (nombre_cliente, moneda), datos in sorted(resumen.items()):
        ws_resumen.append(
            [nombre_cliente, moneda, datos["pagos"], datos["monto"], datos["pendientes"], datos["insoluto"]]
        )
        n = ws_resumen.max_row
        for col in (4, 6):
            ws_resumen.cell(row=n, column=col).number_format = "#,##0.00"

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino