"""Verificación automática del export de conciliación de pagos (checkpoint).

Cruza el Excel generado contra la fuente de verdad (el XML real del REP y
el catálogo), valida los invariantes del Complemento de Pago y escribe una
ficha de verificación en data/salidas/. Uso:

    python scripts/verificar_pagos.py
"""

from __future__ import annotations

import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from conxml.catalog.db import Catalogo
from conxml.catalog.importer import importar_carpeta
from conxml.cfdi.pagos import parse_pagos
from conxml.cfdi.parser import parse_comprobante
from conxml.export.pagos import exportar_pagos

PROYECTO = Path(__file__).resolve().parent.parent
MUESTRA = PROYECTO / "data" / "muestra" / "07. JULIO"
DB = PROYECTO / "data" / "catalogo.db"
SALIDAS = PROYECTO / "data" / "salidas"
PREVIO = PROYECTO / "data" / "comparacion" / "conciliacion_pagos_cliente2.xlsx"
CLIENTE = "2"
PERIODO = "2026-07"

ERRORES: list[str] = []
NOTAS: list[str] = []


def _comprobar(condicion: bool, mensaje: str) -> None:
    if condicion:
        NOTAS.append(f"  OK  {mensaje}")
    else:
        ERRORES.append(f"  FALLA {mensaje}")
        print(f"  FALLA {mensaje}")


def _dec(valor: object) -> Decimal:
    if valor is None:
        return Decimal("0")
    return Decimal(str(valor))


def principal() -> int:
    print(f"== Verificación conciliación de pagos | cliente {CLIENTE} | {PERIODO}")

    xml_rep = MUESTRA / CLIENTE / "aa96caf6-d359-4403-8faf-83c4f73360a0.xml"
    comprobante = parse_comprobante(xml_rep)
    pagos_xml = parse_pagos(comprobante)
    _comprobar(len(pagos_xml) == 1, f"REP parseado: 1 pago en {xml_rep.name}")
    pago_xml = pagos_xml[0]
    _comprobar(
        _dec(pago_xml.monto) == Decimal("900.00"),
        f"Monto del pago según XML: {pago_xml.monto}",
    )
    docto_xml = pago_xml.doctos_relacionados[0]
    _comprobar(docto_xml.uuid == "294624E1-13C2-4282-ABDB-33D80CF19609", "UUID de la factura pagada")
    _comprobar(
        _dec(docto_xml.imp_saldo_ant) == _dec(docto_xml.imp_pagado) + _dec(docto_xml.imp_saldo_insoluto),
        f"Invariante saldo: anterior {docto_xml.imp_saldo_ant} = pagado + insoluto",
    )
    _comprobar(
        _dec(pago_xml.monto) == _dec(docto_xml.imp_pagado),
        f"Invariante monto: pago {pago_xml.monto} = pagado {docto_xml.imp_pagado}",
    )

    with Catalogo(DB) as catalogo:
        imp = importar_carpeta(catalogo, MUESTRA / CLIENTE, CLIENTE)
        _comprobar(
            imp.insertados == 0 and imp.omitidos == 21,
            f"Re-importación idempotente: 0 nuevos, 21 ya existentes, {imp.errores} errores",
        )
        destino = exportar_pagos(
            catalogo, SALIDAS / f"conciliacion_pagos_cliente{CLIENTE}_julio2026.xlsx",
            cliente=CLIENTE,
        )
        _comprobar(destino.is_file(), f"Excel generado: {destino.relative_to(PROYECTO)}")

        pagos_bd = list(catalogo.consultar_pagos(cliente=CLIENTE))
        _comprobar(len(pagos_bd) == 1, f"1 REP en catálogo (cliente {CLIENTE})")
        pago_bd = pagos_bd[0]
        doctos_bd = catalogo.consultar_doctos(pago_bd["id"])
        _comprobar(_dec(pago_bd["monto"]) == _dec(pago_xml.monto), "Monto del pago en catálogo coincide con XML")
        _comprobar(
            _dec(doctos_bd[0]["imp_saldo_insoluto"]) == _dec(docto_xml.imp_saldo_insoluto),
            "Saldo insoluto en catálogo coincide con XML",
        )

    wb = load_workbook(destino, data_only=True)
    ws = wb["Pagos"]
    filas = list(ws.iter_rows(values_only=True))
    enc = [str(c) for c in filas[0]]
    fila = {enc[i]: filas[1][i] for i in range(len(enc))}
    _comprobar(len(filas) == 2, "1 fila de datos en la hoja Pagos")
    _comprobar(fila["UUID REP"] == "AA96CAF6-D359-4403-8FAF-83C4F73360A0", "UUID REP en el Excel")
    _comprobar(_dec(fila["Monto"]) == Decimal("900.00"), "Monto en el Excel")
    _comprobar(_dec(fila["Saldo Insoluto"]) == Decimal("0.00"), "Saldo insoluto en el Excel (liquidado)")
    _comprobar(fila["UUID Factura"] == "294624E1-13C2-4282-ABDB-33D80CF19609", "UUID de factura en el Excel")
    _comprobar(fila["Archivo XML REP"], "Ruta del XML del REP en el Excel")
    _comprobar(fila["FechaTimbrado REP"], "Fecha de timbrado del REP en el Excel")
    _comprobar(
        fila["Estatus Factura"] is None,
        "Factura relacionada sin dato (no está en el catálogo) -> marcada en gris",
    )
    _comprobar(
        _dec(fila["Saldo Anterior"]) == _dec(fila["Pagado"]) + _dec(fila["Saldo Insoluto"]),
        "Invariante saldo en el Excel",
    )

    if PREVIO.is_file():
        prev = load_workbook(PREVIO, data_only=True, read_only=True)
        p_ws = prev[prev.sheetnames[0]]
        prev_filas = [r for r in p_ws.iter_rows(values_only=True)][1:]
        prev_uuid = {str(r[1]).upper() for r in prev_filas if r[1]}
        nuevo_uuid = {str(r[1]).upper() for r in filas[1:] if r[1]}
        _comprobar(prev_uuid == nuevo_uuid, "Mismos UUIDs de REP que el export anterior (comparacion/)")
        prev.close()
    else:
        NOTAS.append(f"  --- Sin export previo que comparar ({PREVIO.name} no existe)")

    ws_resumen = wb["Resumen"]
    resumen = [
        r for r in ws_resumen.iter_rows(values_only=True)
        if r[0] == CLIENTE and r[1] == "MXN"
    ]
    _comprobar(
        resumen and _dec(resumen[0][3]) == Decimal("900.00"),
        "Resumen: cliente 2 MXN total 900.00",
    )

    wb.close()

    linea = "=" * 66
    print(f"\n{linea}")
    print("\n".join(NOTAS))
    print(f"\nTotal: {len(NOTAS)} comprobaciones OK, {len(ERRORES)} fallas")
    ficha = SALIDAS / f"VERIFICACION_PAGOS_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
    SALIDAS.mkdir(parents=True, exist_ok=True)
    ficha.write_text(
        f"Verificación conciliación de pagos (checkpoint)\n"
        f"Fecha: {datetime.now().isoformat(timespec='seconds')}\n"
        f"Cliente: {CLIENTE} | Periodo: {PERIODO}\n"
        f"Excel: {destino}\n"
        f"{linea}\n"
        + "\n".join(NOTAS)
        + f"\n\nRESULTADO: {'APROBADO' if not ERRORES else 'CON FALLAS'} "
        f"({len(NOTAS)} OK, {len(ERRORES)} fallas)\n",
        encoding="utf-8",
    )
    print(f"Ficha escrita: {ficha}")
    print(f"RESULTADO: {'APROBADO' if not ERRORES else 'CON FALLAS'}")
    return 1 if ERRORES else 0


if __name__ == "__main__":
    sys.exit(principal())